# app.py
from flask import Flask, render_template, request, redirect, url_for, session, send_file
import random
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

app = Flask(__name__)
app.secret_key = 'alexander_i_secret_key'

# Путь к файлу Excel
EXCEL_FILE = 'game_results.xlsx'

# Инициализация Excel файла, если его нет
def init_excel_file():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Результаты игр"
        
        # Заголовки
        headers = ['Email', 'Время начала', 'Время окончания', 'Год окончания', 
                  'Армия', 'Дворяне', 'Крестьяне', 'Финансы', 'Средний балл', 'Результат']
        ws.append(headers)
        
        # Стилизация заголовков
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        wb.save(EXCEL_FILE)

# Сохранение результата в Excel
def save_to_excel(email, start_time, end_time, year, stats, result):
    init_excel_file()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    avg_score = sum(stats.values()) / len(stats)
    
    row = [
        email,
        start_time.strftime('%Y-%m-%d %H:%M:%S'),
        end_time.strftime('%Y-%m-%d %H:%M:%S'),
        year,
        stats.get('Армия', 0),
        stats.get('Дворяне', 0),
        stats.get('Крестьяне', 0),
        stats.get('Финансы', 0),
        round(avg_score, 2),
        result
    ]
    
    ws.append(row)
    wb.save(EXCEL_FILE)

# Карточки событий для Александра I
# Фракции: Армия, Дворяне, Крестьяне, Финансы
EVENTS = [
    {
        'title': 'Реформа Сперанского',
        'text': 'Михаил Сперанский предлагает проект реформы государственного управления. Дворяне против, но это улучшит эффективность.',
        'yes': 'Поддержать реформу',
        'no': 'Отклонить предложение',
        'yes_effect': {'Дворяне': -15, 'Финансы': 10, 'Крестьяне': 5},
        'no_effect': {'Дворяне': 15, 'Финансы': -10, 'Крестьяне': -5},
        'year': 1809
    },
    {
        'title': 'Война с Наполеоном',
        'text': 'Франция требует присоединения к континентальной блокаде. Это может спровоцировать войну.',
        'yes': 'Присоединиться к блокаде',
        'no': 'Отказаться от блокады',
        'yes_effect': {'Армия': -10, 'Дворяне': 5, 'Финансы': -15},
        'no_effect': {'Армия': 10, 'Дворяне': -5, 'Финансы': 15},
        'year': 1812
    },
    {
        'title': 'Создание министерств',
        'text': 'Предложение о создании системы министерств для улучшения управления государством.',
        'yes': 'Создать министерства',
        'no': 'Оставить старую систему',
        'yes_effect': {'Дворяне': -10, 'Финансы': 10, 'Крестьяне': 5},
        'no_effect': {'Дворяне': 15, 'Финансы': -10, 'Крестьяне': -5},
        'year': 1802
    },
    {
        'title': 'Университетская реформа',
        'text': 'Предложение об открытии новых университетов и расширении доступа к образованию.',
        'yes': 'Поддержать образование',
        'no': 'Сократить расходы',
        'yes_effect': {'Финансы': -15, 'Крестьяне': 15, 'Дворяне': -5},
        'no_effect': {'Финансы': 10, 'Крестьяне': -15, 'Дворяне': 10},
        'year': 1804
    },
    {
        'title': 'Крестьянский вопрос',
        'text': 'Проект облегчения положения крестьян. Помещики крайне недовольны.',
        'yes': 'Поддержать крестьян',
        'no': 'Защитить интересы помещиков',
        'yes_effect': {'Крестьяне': 25, 'Дворяне': -20, 'Армия': -5},
        'no_effect': {'Крестьяне': -20, 'Дворяне': 20, 'Армия': 5},
        'year': 1803
    },
    {
        'title': 'Священный союз',
        'text': 'Предложение создать Священный союз с другими европейскими монархами для поддержания порядка.',
        'yes': 'Присоединиться к союзу',
        'no': 'Отклонить предложение',
        'yes_effect': {'Дворяне': 10, 'Армия': 10, 'Финансы': -5},
        'no_effect': {'Дворяне': -10, 'Армия': -10, 'Финансы': 5},
        'year': 1815
    },
    {
        'title': 'Военные поселения',
        'text': 'Аракчеев предлагает создать военные поселения для снижения расходов на армию.',
        'yes': 'Внедрить систему',
        'no': 'Отказаться от идеи',
        'yes_effect': {'Армия': 15, 'Финансы': 10, 'Крестьяне': -20},
        'no_effect': {'Армия': -10, 'Финансы': -15, 'Крестьяне': 15},
        'year': 1816
    },
    {
        'title': 'Торговый договор с Англией',
        'text': 'Предложение о заключении выгодного торгового договора с Великобританией.',
        'yes': 'Заключить договор',
        'no': 'Отклонить предложение',
        'yes_effect': {'Финансы': 20, 'Дворяне': 5, 'Армия': -5},
        'no_effect': {'Финансы': -15, 'Дворяне': -5, 'Армия': 5},
        'year': 1807
    },
    {
        'title': 'Цензура и свобода слова',
        'text': 'Министры требуют ужесточения цензуры для предотвращения революционных настроений.',
        'yes': 'Усилить цензуру',
        'no': 'Сохранить свободу',
        'yes_effect': {'Дворяне': 15, 'Крестьяне': -15, 'Армия': 5},
        'no_effect': {'Дворяне': -15, 'Крестьяне': 15, 'Армия': -5},
        'year': 1810
    },
    {
        'title': 'Реформа финансовой системы',
        'text': 'Необходимо провести реформу финансовой системы для стабилизации казны после войн.',
        'yes': 'Провести реформу',
        'no': 'Отложить реформу',
        'yes_effect': {'Финансы': 20, 'Дворяне': -10, 'Крестьяне': -5},
        'no_effect': {'Финансы': -20, 'Дворяне': 10, 'Крестьяне': 5},
        'year': 1810
    },
    {
        'title': 'Отношения с Пруссией',
        'text': 'Пруссия просит помощи в борьбе с Наполеоном. Это может привести к войне.',
        'yes': 'Поддержать Пруссию',
        'no': 'Сохранить нейтралитет',
        'yes_effect': {'Армия': -15, 'Дворяне': 10, 'Финансы': -10},
        'no_effect': {'Армия': 10, 'Дворяне': -5, 'Финансы': 10},
        'year': 1806
    },
    {
        'title': 'Реформа судебной системы',
        'text': 'Предложение о создании независимых судов и введении суда присяжных.',
        'yes': 'Провести судебную реформу',
        'no': 'Оставить старую систему',
        'yes_effect': {'Крестьяне': 20, 'Дворяне': -15, 'Финансы': -5},
        'no_effect': {'Крестьяне': -15, 'Дворяне': 20, 'Финансы': 5},
        'year': 1808
    },
    {
        'title': 'Расходы на двор',
        'text': 'Двор требует увеличения финансирования на роскошь и церемонии.',
        'yes': 'Увеличить расходы',
        'no': 'Сократить траты',
        'yes_effect': {'Дворяне': 20, 'Финансы': -25, 'Крестьяне': -10},
        'no_effect': {'Дворяне': -20, 'Финансы': 20, 'Крестьяне': 10},
        'year': 1812
    },
    {
        'title': 'Реформа крепостного права',
        'text': 'Проект постепенного освобождения крестьян от крепостной зависимости.',
        'yes': 'Начать реформу',
        'no': 'Отложить вопрос',
        'yes_effect': {'Крестьяне': 30, 'Дворяне': -25, 'Армия': -10},
        'no_effect': {'Крестьяне': -20, 'Дворяне': 25, 'Армия': 10},
        'year': 1803
    },
    {
        'title': 'Медицинская реформа',
        'text': 'Предложение о создании системы государственных больниц и улучшении здравоохранения.',
        'yes': 'Поддержать медицину',
        'no': 'Сократить расходы',
        'yes_effect': {'Крестьяне': 20, 'Финансы': -15, 'Дворяне': -5},
        'no_effect': {'Крестьяне': -15, 'Финансы': 15, 'Дворяне': 10},
        'year': 1805
    },
    {
        'title': 'Военная реформа',
        'text': 'Предложение о модернизации армии и улучшении военной подготовки.',
        'yes': 'Провести реформу',
        'no': 'Оставить как есть',
        'yes_effect': {'Армия': 20, 'Финансы': -20, 'Дворяне': 5},
        'no_effect': {'Армия': -15, 'Финансы': 15, 'Дворяне': -5},
        'year': 1804
    },
    {
        'title': 'Налоговая реформа',
        'text': 'Предложение о пересмотре налоговой системы для увеличения доходов казны.',
        'yes': 'Провести реформу',
        'no': 'Оставить текущую систему',
        'yes_effect': {'Финансы': 15, 'Крестьяне': -15, 'Дворяне': -10},
        'no_effect': {'Финансы': -10, 'Крестьяне': 10, 'Дворяне': 15},
        'year': 1811
    },
    {
        'title': 'Привилегии дворянства',
        'text': 'Дворяне требуют расширения своих привилегий и прав.',
        'yes': 'Увеличить привилегии',
        'no': 'Отклонить требования',
        'yes_effect': {'Дворяне': 20, 'Крестьяне': -20, 'Финансы': -5},
        'no_effect': {'Дворяне': -20, 'Крестьяне': 15, 'Финансы': 5},
        'year': 1807
    },
    {
        'title': 'Крестьянские волнения',
        'text': 'Вспыхивают крестьянские волнения из-за тяжелых условий жизни.',
        'yes': 'Подавить силой',
        'no': 'Пойти на уступки',
        'yes_effect': {'Армия': 10, 'Крестьяне': -25, 'Дворяне': 10},
        'no_effect': {'Армия': -10, 'Крестьяне': 20, 'Дворяне': -15},
        'year': 1813
    }
]

@app.route('/')
def index():
    session.clear()
    session['year'] = 1801
    # 4 фракции: Армия, Дворяне, Крестьяне, Финансы
    session['stats'] = {
        'Армия': 50,
        'Дворяне': 50,
        'Крестьяне': 50,
        'Финансы': 50
    }
    session['events_seen'] = []
    session['game_over'] = False
    session['history'] = []
    session['email'] = None
    session['start_time'] = None
    
    return render_template('index.html', 
                         stats=session['stats'], 
                         year=session['year'],
                         game_over=False)

@app.route('/next_card', methods=['POST'])
def next_card():
    if session.get('game_over'):
        return redirect(url_for('index'))
    
    # Если это первая карточка, сохраняем email и время начала
    if not session.get('start_time'):
        email = request.form.get('email')
        if email:
            session['email'] = email
            session['start_time'] = datetime.now()
    
    # Получаем решение игрока
    decision = request.form.get('decision')
    
    if decision:
        current_event = session.get('current_event')
        if current_event:
            # Применяем эффекты решения
            if decision == 'yes':
                for stat, value in current_event['yes_effect'].items():
                    if stat in session['stats']:  # Проверяем, существует ли параметр
                        session['stats'][stat] = max(0, min(100, session['stats'][stat] + value))
            else:
                for stat, value in current_event['no_effect'].items():
                    if stat in session['stats']:  # Проверяем, существует ли параметр
                        session['stats'][stat] = max(0, min(100, session['stats'][stat] + value))
            
            # Добавляем в историю
            session['history'].append({
                'year': session['year'],
                'event': current_event['title'],
                'decision': 'Да' if decision == 'yes' else 'Нет'
            })
    
    # Проверяем условия окончания игры (если любая фракция достигла 0 или 100)
    if any(value <= 0 or value >= 100 for value in session['stats'].values()):
        session['game_over'] = True
        return redirect(url_for('game_over'))
    
    # Переходим к следующему году
    session['year'] += 1
    
    # Выбираем новую карточку
    available_events = [e for e in EVENTS if e not in session['events_seen'] and e['year'] <= session['year']]
    if not available_events:
        available_events = [e for e in EVENTS if e['year'] <= session['year']]
        session['events_seen'] = []
    
    if available_events:
        current_event = random.choice(available_events)
        session['current_event'] = current_event
        session['events_seen'].append(current_event)
    else:
        session['game_over'] = True
        return redirect(url_for('game_over'))
    
    return render_template('card.html', 
                         event=current_event, 
                         stats=session['stats'], 
                         year=session['year'])

@app.route('/game_over')
def game_over():
    stats = session.get('stats', {})
    history = session.get('history', [])
    year = session.get('year', 1801)
    email = session.get('email', 'Не указана')
    start_time = session.get('start_time')
    end_time = datetime.now()
    
    # Определяем результат правления
    avg_score = sum(stats.values()) / len(stats)
    
    if avg_score >= 70:
        result = "Великое правление! Александр I вошел в историю как мудрый реформатор."
    elif avg_score >= 50:
        result = "Умеренно успешное правление. Были достигнуты некоторые успехи."
    elif avg_score >= 30:
        result = "Трудное правление. Много проблем осталось нерешенными."
    else:
        result = "Неудачное правление. Страна оказалась в кризисе."
    
    # Сохраняем результат в Excel
    if start_time:
        save_to_excel(email, start_time, end_time, year, stats, result)
    
    return render_template('game_over.html', 
                         stats=stats, 
                         history=history, 
                         year=year, 
                         result=result)

@app.route('/export')
def export_data():
    """Эндпоинт для выгрузки данных из Excel"""
    if not os.path.exists(EXCEL_FILE):
        init_excel_file()
    
    return send_file(EXCEL_FILE, 
                    as_attachment=True, 
                    download_name='game_results.xlsx',
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)